import pandas as pd 
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

# Styling constants
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
blue_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

bold = Font(bold=True)
white = Font(color="FFFFFF")

b_side = Side(border_style='thin', color='000000')
b_border = Border(left=b_side, right=b_side, top=b_side, bottom=b_side)


def load_and_prepare_prework_data(file_path):
    col_keep = [0, 2, 3]
    df = pd.read_excel(file_path, sheet_name=0, header=None, usecols=col_keep)
    
    # Add headers
    headers = ['Building Code', 'Old Mac', 'New Mac']
    df.columns = headers
    
    # Keep last 4 characters from MAC addresses
    df['Old Mac'] = df['Old Mac'].astype(str).str[-4:]
    df['New Mac'] = df['New Mac'].astype(str).str[-4:]
    
    return df


def load_and_prepare_tech_data(file_path):
    col_keep = [1, 5, 12, 18, 23]
    headers = ['Building Code', 'Install Date', 'Partner', 'Bridge', 'Tech']
    
    df = pd.read_excel(file_path, sheet_name=1, header=None, usecols=col_keep)
    df.columns = headers
    
    # Extract last character from bridge
    df['Bridge'] = df['Bridge'].astype(str).str[-1]
    
    # Filter for WWT FS only
    df = df[df['Partner'] == 'WWT FS']
    
    return df


def merge_and_format_data(prework_df, tech_df):
    merged = pd.merge(prework_df, tech_df, on='Building Code', how='inner')
    merged['Install Date'] = pd.to_datetime(merged['Install Date']).dt.strftime('%m-%d')
    merged['Description'] = ' '
    
    return merged


def write_site_header(sheet, row, building_code, bridge_num):
    #Site #: Bridge #
    cell = sheet.cell(row=row, column=1, value=f"Site {building_code.replace('B', '')} : Bridge {bridge_num}")
    cell.font = bold
    return row + 1


def write_tech_name_row(sheet, row, tech_name, building_code, install_date):
    # Tech name and MM:Swap
    name_cell = sheet.cell(row=row, column=1, value=f"{tech_name} ({building_code}) MM:Swap")
    name_cell.font = Font(bold=True, color='FFFFFF')
    name_cell.fill = red_fill
    name_cell.border = Border(left=b_side, top=b_side)
    # Middle empty cells with red fill
    for col in [2, 3]:
        cell = sheet.cell(row=row, column=col, value="")
        cell.fill = red_fill
        cell.border = Border(top=b_side)
    
    # Date cell
    date_cell = sheet.cell(row=row, column=4, value=install_date)
    date_cell.font = Font(bold=True, color='FFFFFF')
    date_cell.fill = red_fill
    date_cell.border = Border(top=b_side, right=b_side)
    
    return row + 1


def write_column_headers(sheet, row):
    headers = ['Site ID', 'OLD MAC', 'NEW MAC', 'Description']
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=header)
        cell.fill = gray_fill
        cell.border = b_border
        cell.font = bold
    return row + 1


def write_data_rows(sheet, row, building_data):
    for _, row_data in building_data.iterrows():
        # Building Code
        cell1 = sheet.cell(row=row, column=1, value=row_data['Building Code'])
        cell1.fill = blue_fill
        cell1.border = b_border
        
        # Old Mac
        cell2 = sheet.cell(row=row, column=2, value=row_data['Old Mac'])
        cell2.fill = blue_fill
        cell2.border = b_border
        
        # New Mac
        cell3 = sheet.cell(row=row, column=3, value=row_data['New Mac'])
        cell3.fill = blue_fill
        cell3.border = b_border
        
        # Description
        cell4 = sheet.cell(row=row, column=4, value=row_data['Description'])
        cell4.fill = blue_fill
        cell4.border = b_border
        
        row += 1
    
    return row


def write_tech_tables(sheet, merged_data):
    grouped = merged_data.groupby('Tech')
    current_row = 1
    
    for tech_name, tech_data in grouped:
        building_groups = tech_data.groupby('Building Code')
        
        for building_code, building_data in building_groups:
            # Get bridge number and install date
            bridge_num = building_data['Bridge'].iloc[0]
            install_date = building_data['Install Date'].iloc[0]
            current_row = write_site_header(sheet, current_row, building_code, bridge_num)
            current_row = write_tech_name_row(sheet, current_row, tech_name, building_code, install_date)
            current_row = write_column_headers(sheet, current_row)
            current_row = write_data_rows(sheet, current_row, building_data)
            current_row += 2


def main():
    file_path = "file.xlsx"
    # Load and prepare data
    prework_df = load_and_prepare_prework_data(file_path)
    tech_df = load_and_prepare_tech_data(file_path)
    
    # Merge and format
    merged_data = merge_and_format_data(prework_df, tech_df)
    
    # Load workbook and prepare output sheet
    wb = openpyxl.load_workbook(file_path)
    output_sheet = wb.worksheets[4]
    output_sheet.title = "Output tables"
    
    # Clear existing data
    output_sheet.delete_rows(1, output_sheet.max_row)
    
    # Write all tech tables
    write_tech_tables(output_sheet, merged_data)
    
    # Save workbook
    wb.save(file_path)
    print("Data processed and saved in new tab on workbook.")


if __name__ == "__main__":
    main()
